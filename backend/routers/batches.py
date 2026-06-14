"""批次路由模块。"""

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

import models
import schemas
from database import get_db, write_change_log

router = APIRouter(prefix="/api/batches", tags=["批次"])


@router.get("", response_model=list[schemas.BatchOut])
def list_batches(
    status: str | None = Query(None, description="按发酵状态筛选"),
    type: str | None = Query(None, description="按批次类型筛选"),
    search: str | None = Query(None, description="按批次类型关键字模糊搜索"),
    db: Session = Depends(get_db),
):
    """获取全部批次，可按状态、类型筛选，或按类型关键字模糊搜索。"""
    query = db.query(models.Batch)
    if status:
        query = query.filter(models.Batch.status == status)
    if type:
        query = query.filter(models.Batch.type == type)
    if search:
        query = query.filter(models.Batch.type.contains(search))
    return query.order_by(models.Batch.start_date.desc()).all()


@router.post("", response_model=schemas.BatchOut, status_code=201)
def create_batch(payload: schemas.BatchCreate, db: Session = Depends(get_db)):
    """创建批次。"""
    batch = models.Batch(**payload.model_dump())
    db.add(batch)
    db.commit()
    db.refresh(batch)
    write_change_log(db, "create", "batch", batch.id, f"创建批次：{batch.type}，状态：{batch.status}")
    db.commit()
    return batch


@router.get("/export")
def export_batches(db: Session = Depends(get_db)):
    """导出全部批次及关联笔记为 Excel 文件。"""
    wb = Workbook()

    ws_batches = wb.active
    ws_batches.title = "批次"
    batch_headers = ["编号", "类型", "开始日期", "发酵天数", "温度(°C)", "状态", "pH", "创建时间"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(batch_headers, 1):
        cell = ws_batches.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    batches = db.query(models.Batch).order_by(models.Batch.id.asc()).all()
    for row_idx, batch in enumerate(batches, 2):
        ws_batches.cell(row=row_idx, column=1, value=batch.id)
        ws_batches.cell(row=row_idx, column=2, value=batch.type)
        ws_batches.cell(row=row_idx, column=3, value=batch.start_date.isoformat())
        ws_batches.cell(row=row_idx, column=4, value=batch.fermentation_days)
        ws_batches.cell(row=row_idx, column=5, value=batch.temperature)
        ws_batches.cell(row=row_idx, column=6, value=batch.status)
        ws_batches.cell(row=row_idx, column=7, value=batch.ph if batch.ph is not None else "")
        ws_batches.cell(row=row_idx, column=8, value=batch.created_at.strftime("%Y-%m-%d %H:%M:%S"))

    batch_col_widths = [8, 16, 14, 12, 12, 12, 10, 22]
    for i, width in enumerate(batch_col_widths, 1):
        ws_batches.column_dimensions[get_column_letter(i)].width = width

    ws_notes = wb.create_sheet(title="笔记")
    note_headers = ["笔记编号", "批次编号", "笔记内容", "创建时间"]
    for col_idx, header in enumerate(note_headers, 1):
        cell = ws_notes.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    notes = db.query(models.Note).order_by(models.Note.id.asc()).all()
    for row_idx, note in enumerate(notes, 2):
        ws_notes.cell(row=row_idx, column=1, value=note.id)
        ws_notes.cell(row=row_idx, column=2, value=note.batch_id)
        ws_notes.cell(row=row_idx, column=3, value=note.content)
        ws_notes.cell(row=row_idx, column=4, value=note.created_at.strftime("%Y-%m-%d %H:%M:%S"))

    note_col_widths = [10, 10, 60, 22]
    for i, width in enumerate(note_col_widths, 1):
        ws_notes.column_dimensions[get_column_letter(i)].width = width

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    filename = f"ferment_batches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        virtual_workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import", response_model=schemas.ImportResult)
def import_batches(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """从 Excel 文件导入批次和笔记，跳过重复编号。"""
    if not (file.filename and file.filename.endswith((".xlsx", ".xls"))):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式的 Excel 文件")

    try:
        contents = file.file.read()
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请确认文件格式正确")

    if "批次" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Excel 文件中缺少「批次」工作表")

    ws_batches = wb["批次"]
    existing_batch_ids = {row[0] for row in db.query(models.Batch.id).all()}

    inserted_batches = 0
    skipped_batches = 0
    total_batches_in_file = 0
    batch_id_mapping: dict[int, int] = {}

    batch_rows = list(ws_batches.iter_rows(min_row=2, values_only=True))
    for row in batch_rows:
        if row[0] is None:
            continue
        total_batches_in_file += 1
        original_id = int(row[0])

        if original_id in existing_batch_ids:
            skipped_batches += 1
            continue

        try:
            batch = models.Batch(
                id=original_id,
                type=str(row[1]) if row[1] else "",
                start_date=datetime.strptime(str(row[2]), "%Y-%m-%d").date(),
                temperature=float(row[3]) if row[3] is not None else 0,
                status=str(row[4]) if row[4] else "",
                ph=float(row[5]) if row[5] not in (None, "") else None,
                created_at=(
                    datetime.strptime(str(row[6]), "%Y-%m-%d %H:%M:%S")
                    if row[6]
                    else datetime.utcnow()
                ),
            )
            db.add(batch)
            db.flush()
            batch_id_mapping[original_id] = batch.id
            existing_batch_ids.add(batch.id)
            inserted_batches += 1
        except Exception as exc:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail=f"解析第 {total_batches_in_file + 1} 行批次数据出错: {exc}",
            )

    inserted_notes = 0
    skipped_notes = 0
    total_notes_in_file = 0

    if "笔记" in wb.sheetnames:
        ws_notes = wb["笔记"]
        existing_note_ids = {row[0] for row in db.query(models.Note.id).all()}
        note_rows = list(ws_notes.iter_rows(min_row=2, values_only=True))

        for row in note_rows:
            if row[0] is None:
                continue
            total_notes_in_file += 1
            note_original_id = int(row[0])
            batch_original_id = int(row[1]) if row[1] is not None else None

            if batch_original_id is None:
                skipped_notes += 1
                continue

            if batch_original_id not in batch_id_mapping and batch_original_id not in existing_batch_ids:
                skipped_notes += 1
                continue

            if note_original_id in existing_note_ids:
                skipped_notes += 1
                continue

            actual_batch_id = batch_id_mapping.get(batch_original_id, batch_original_id)
            try:
                note = models.Note(
                    id=note_original_id,
                    batch_id=actual_batch_id,
                    content=str(row[2]) if row[2] else "",
                    created_at=(
                        datetime.strptime(str(row[3]), "%Y-%m-%d %H:%M:%S")
                        if row[3]
                        else datetime.utcnow()
                    ),
                )
                db.add(note)
                existing_note_ids.add(note.id)
                inserted_notes += 1
            except Exception as exc:
                db.rollback()
                raise HTTPException(
                    status_code=400,
                    detail=f"解析第 {total_notes_in_file + 1} 行笔记数据出错: {exc}",
                )

    db.commit()

    return schemas.ImportResult(
        inserted_batches=inserted_batches,
        skipped_batches=skipped_batches,
        inserted_notes=inserted_notes,
        skipped_notes=skipped_notes,
        total_batches_in_file=total_batches_in_file,
        total_notes_in_file=total_notes_in_file,
    )


@router.get("/{batch_id}", response_model=schemas.BatchDetail)
def get_batch(batch_id: int, db: Session = Depends(get_db)):
    """获取批次详情（含笔记）。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")
    return batch


@router.put("/{batch_id}", response_model=schemas.BatchOut)
def update_batch(
    batch_id: int,
    payload: schemas.BatchUpdate,
    db: Session = Depends(get_db),
):
    """更新批次。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")

    changes = payload.model_dump(exclude_unset=True)
    change_parts = [f"{k}={v}" for k, v in changes.items()]
    summary = f"更新批次 #{batch_id}：{', '.join(change_parts)}"

    for key, value in changes.items():
        setattr(batch, key, value)

    db.commit()
    db.refresh(batch)
    write_change_log(db, "update", "batch", batch_id, summary)
    db.commit()
    return batch


@router.delete("/{batch_id}", status_code=204)
def delete_batch(batch_id: int, db: Session = Depends(get_db)):
    """删除批次。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")
    batch_type = batch.type
    db.delete(batch)
    db.commit()
    write_change_log(db, "delete", "batch", batch_id, f"删除批次：{batch_type}")
    db.commit()
