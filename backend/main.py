"""家庭发酵实验日志 — FastAPI 入口。"""

from datetime import datetime, timedelta
from io import BytesIO

from fastapi import Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

import models
import schemas
from database import Base, engine, get_db
from seed import seed_data

Base.metadata.create_all(bind=engine)

app = FastAPI(title="家庭发酵实验日志 API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5101",
        "http://127.0.0.1:5101",
        "http://localhost:5102",
        "http://127.0.0.1:5102",
        "http://localhost:5103",
        "http://127.0.0.1:5103",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    """启动时写入种子数据。"""
    db = next(get_db())
    try:
        seed_data(db)
    finally:
        db.close()


@app.get("/api/health")
def health():
    """健康检查。"""
    return {"status": "ok"}


@app.get("/api/batches", response_model=list[schemas.BatchOut])
def list_batches(db: Session = Depends(get_db)):
    """获取全部批次。"""
    return db.query(models.Batch).order_by(models.Batch.start_date.desc()).all()


@app.post("/api/batches", response_model=schemas.BatchOut, status_code=201)
def create_batch(payload: schemas.BatchCreate, db: Session = Depends(get_db)):
    """创建批次。"""
    batch = models.Batch(**payload.model_dump())
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return batch


@app.get("/api/batches/export")
def export_batches(db: Session = Depends(get_db)):
    """导出全部批次及关联笔记为 Excel 文件。"""
    wb = Workbook()

    ws_batches = wb.active
    ws_batches.title = "批次"
    batch_headers = ["编号", "类型", "开始日期", "温度(°C)", "状态", "pH", "创建时间"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(batch_headers, 1):
        cell = ws_batches.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    batches = db.query(models.Batch).order_by(models.Batch.id.asc()).all()
    for row_idx, batch in enumerate(batches, 2):
        ws_batches.cell(row=row_idx, column=1, value=batch.id)
        ws_batches.cell(row=row_idx, column=2, value=batch.type)
        ws_batches.cell(row=row_idx, column=3, value=batch.start_date.isoformat())
        ws_batches.cell(row=row_idx, column=4, value=batch.temperature)
        ws_batches.cell(row=row_idx, column=5, value=batch.status)
        ws_batches.cell(row=row_idx, column=6, value=batch.ph if batch.ph is not None else "")
        ws_batches.cell(row=row_idx, column=7, value=batch.created_at.strftime("%Y-%m-%d %H:%M:%S"))

    batch_col_widths = [8, 16, 14, 12, 12, 10, 22]
    for i, width in enumerate(batch_col_widths, 1):
        ws_batches.column_dimensions[get_column_letter(i)].width = width

    ws_notes = wb.create_sheet(title="笔记")
    note_headers = ["笔记编号", "批次编号", "笔记内容", "创建时间"]
    for col_idx, header in enumerate(note_headers, 1):
        cell = ws_notes.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    notes = db.query(models.Note).order_by(models.Note.id.asc()).all()
    for row_idx, note in enumerate(notes, 2):
        ws_notes.cell(row=row_idx, column=1, value=note.id)
        ws_notes.cell(row=row_idx, column=2, value=note.batch_id)
        ws_notes.cell(row=row_idx, column=3, value=note.content)
        ws_notes.cell(row=row_idx, column=4, value=note.created_at.strftime("%Y-%m-%d %H:%M:%S"))

    note_col_widths = [10, 10, 60, 22]
    for i, width in enumerate(note_col_widths, 1):
        ws_notes.column_dimensions[get_column_letter(i)].width = width

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)

    filename = f"ferment_batches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        virtual_workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/batches/import", response_model=schemas.ImportResult)
def import_batches(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """从 Excel 文件导入批次和笔记，跳过重复编号。"""
    if not (file.filename and file.filename.endswith((".xlsx", ".xls"))):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式的 Excel 文件")

    try:
        contents = file.file.read()
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请确认文件格式正确")

    if "批次" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Excel 文件中缺少「批次」工作表")

    ws_batches = wb["批次"]
    existing_batch_ids = {row[0] for row in db.query(models.Batch.id).all()}

    inserted_batches = 0
    skipped_batches = 0
    total_batches_in_file = 0
    batch_id_mapping: dict[int, int] = {}

    batch_rows = list(ws_batches.iter_rows(min_row=2, values_only=True))
    for row in batch_rows:
        if row[0] is None:
            continue
        total_batches_in_file += 1
        original_id = int(row[0])

        if original_id in existing_batch_ids:
            skipped_batches += 1
            continue

        try:
            batch = models.Batch(
                id=original_id,
                type=str(row[1]) if row[1] else "",
                start_date=datetime.strptime(str(row[2]), "%Y-%m-%d").date(),
                temperature=float(row[3]) if row[3] is not None else 0,
                status=str(row[4]) if row[4] else "",
                ph=float(row[5]) if row[5] not in (None, "") else None,
                created_at=(
                    datetime.strptime(str(row[6]), "%Y-%m-%d %H:%M:%S")
                    if row[6]
                    else datetime.utcnow()
                ),
            )
            db.add(batch)
            db.flush()
            batch_id_mapping[original_id] = batch.id
            existing_batch_ids.add(batch.id)
            inserted_batches += 1
        except Exception as exc:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail=f"解析第 {total_batches_in_file + 1} 行批次数据出错: {exc}",
            )

    inserted_notes = 0
    skipped_notes = 0
    total_notes_in_file = 0

    if "笔记" in wb.sheetnames:
        ws_notes = wb["笔记"]
        existing_note_ids = {row[0] for row in db.query(models.Note.id).all()}
        note_rows = list(ws_notes.iter_rows(min_row=2, values_only=True))

        for row in note_rows:
            if row[0] is None:
                continue
            total_notes_in_file += 1
            note_original_id = int(row[0])
            batch_original_id = int(row[1]) if row[1] is not None else None

            if batch_original_id is None:
                skipped_notes += 1
                continue

            if batch_original_id not in batch_id_mapping and batch_original_id not in existing_batch_ids:
                skipped_notes += 1
                continue

            if note_original_id in existing_note_ids:
                skipped_notes += 1
                continue

            actual_batch_id = batch_id_mapping.get(batch_original_id, batch_original_id)
            try:
                note = models.Note(
                    id=note_original_id,
                    batch_id=actual_batch_id,
                    content=str(row[2]) if row[2] else "",
                    created_at=(
                        datetime.strptime(str(row[3]), "%Y-%m-%d %H:%M:%S")
                        if row[3]
                        else datetime.utcnow()
                    ),
                )
                db.add(note)
                existing_note_ids.add(note.id)
                inserted_notes += 1
            except Exception as exc:
                db.rollback()
                raise HTTPException(
                    status_code=400,
                    detail=f"解析第 {total_notes_in_file + 1} 行笔记数据出错: {exc}",
                )

    db.commit()

    return schemas.ImportResult(
        inserted_batches=inserted_batches,
        skipped_batches=skipped_batches,
        inserted_notes=inserted_notes,
        skipped_notes=skipped_notes,
        total_batches_in_file=total_batches_in_file,
        total_notes_in_file=total_notes_in_file,
    )


@app.get("/api/batches/{batch_id}", response_model=schemas.BatchDetail)
def get_batch(batch_id: int, db: Session = Depends(get_db)):
    """获取批次详情（含笔记）。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")
    return batch


@app.put("/api/batches/{batch_id}", response_model=schemas.BatchOut)
def update_batch(
    batch_id: int,
    payload: schemas.BatchUpdate,
    db: Session = Depends(get_db),
):
    """更新批次。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")

    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(batch, key, value)

    db.commit()
    db.refresh(batch)
    return batch


@app.delete("/api/batches/{batch_id}", status_code=204)
def delete_batch(batch_id: int, db: Session = Depends(get_db)):
    """删除批次。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")
    db.delete(batch)
    db.commit()


@app.post(
    "/api/batches/{batch_id}/notes",
    response_model=schemas.NoteOut,
    status_code=201,
)
def create_note(
    batch_id: int,
    payload: schemas.NoteCreate,
    db: Session = Depends(get_db),
):
    """为批次追加观察笔记。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")

    note = models.Note(batch_id=batch_id, content=payload.content)
    db.add(note)
    db.commit()
    db.refresh(note)
    return note


@app.delete("/api/notes/{note_id}", status_code=204)
def delete_note(note_id: int, db: Session = Depends(get_db)):
    """删除笔记。"""
    note = db.query(models.Note).filter(models.Note.id == note_id).first()
    if not note:
        raise HTTPException(status_code=404, detail="笔记不存在")
    db.delete(note)
    db.commit()


@app.get(
    "/api/batches/{batch_id}/measurements",
    response_model=list[schemas.MeasurementOut],
)
def list_measurements(batch_id: int, db: Session = Depends(get_db)):
    """按批次查询全部测量记录。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")
    return (
        db.query(models.Measurement)
        .filter(models.Measurement.batch_id == batch_id)
        .order_by(models.Measurement.recorded_at.desc())
        .all()
    )


@app.post(
    "/api/batches/{batch_id}/measurements",
    response_model=schemas.MeasurementOut,
    status_code=201,
)
def create_measurement(
    batch_id: int,
    payload: schemas.MeasurementCreate,
    db: Session = Depends(get_db),
):
    """为批次新增单条测量记录。"""
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")

    measurement = models.Measurement(
        batch_id=batch_id,
        recorded_at=payload.recorded_at,
        temperature=payload.temperature,
        ph=payload.ph,
    )
    db.add(measurement)
    db.commit()
    db.refresh(measurement)
    return measurement


@app.get("/api/recipes", response_model=list[schemas.RecipeOut])
def list_recipes(db: Session = Depends(get_db)):
    """获取全部配方。"""
    return db.query(models.Recipe).order_by(models.Recipe.created_at.desc()).all()


@app.post("/api/recipes", response_model=schemas.RecipeDetail, status_code=201)
def create_recipe(payload: schemas.RecipeCreate, db: Session = Depends(get_db)):
    """创建配方。"""
    recipe = models.Recipe(
        name=payload.name,
        ferment_type=payload.ferment_type,
        ingredients=payload.ingredients,
    )
    db.add(recipe)
    db.flush()

    for step_data in payload.steps:
        step = models.RecipeStep(
            recipe_id=recipe.id,
            step_order=step_data.step_order,
            description=step_data.description,
        )
        db.add(step)

    db.commit()
    db.refresh(recipe)
    return recipe


@app.get("/api/recipes/{recipe_id}", response_model=schemas.RecipeDetail)
def get_recipe(recipe_id: int, db: Session = Depends(get_db)):
    """获取配方详情（含步骤）。"""
    recipe = db.query(models.Recipe).filter(models.Recipe.id == recipe_id).first()
    if not recipe:
        raise HTTPException(status_code=404, detail="配方不存在")
    return recipe


@app.put("/api/recipes/{recipe_id}", response_model=schemas.RecipeDetail)
def update_recipe(
    recipe_id: int,
    payload: schemas.RecipeUpdate,
    db: Session = Depends(get_db),
):
    """更新配方。"""
    recipe = db.query(models.Recipe).filter(models.Recipe.id == recipe_id).first()
    if not recipe:
        raise HTTPException(status_code=404, detail="配方不存在")

    update_data = payload.model_dump(exclude_unset=True)
    steps_data = update_data.pop("steps", None)

    for key, value in update_data.items():
        setattr(recipe, key, value)

    if steps_data is not None:
        db.query(models.RecipeStep).filter(
            models.RecipeStep.recipe_id == recipe_id
        ).delete()

        for step_data in steps_data:
            step = models.RecipeStep(
                recipe_id=recipe.id,
                step_order=step_data["step_order"],
                description=step_data["description"],
            )
            db.add(step)

    db.commit()
    db.refresh(recipe)
    return recipe


@app.delete("/api/recipes/{recipe_id}", status_code=204)
def delete_recipe(recipe_id: int, db: Session = Depends(get_db)):
    """删除配方。"""
    recipe = db.query(models.Recipe).filter(models.Recipe.id == recipe_id).first()
    if not recipe:
        raise HTTPException(status_code=404, detail="配方不存在")
    db.delete(recipe)
    db.commit()


@app.get("/api/statistics", response_model=schemas.StatisticsOut)
def get_statistics(db: Session = Depends(get_db)):
    """获取数据统计概览。"""
    status_counts = dict(
        db.query(models.Batch.status, func.count(models.Batch.id))
        .group_by(models.Batch.status)
        .all()
    )

    type_counts = dict(
        db.query(models.Batch.type, func.count(models.Batch.id))
        .group_by(models.Batch.type)
        .all()
    )

    now_local = datetime.now()
    seven_days_ago_date = now_local.date() - timedelta(days=7)
    seven_days_ago = datetime.combine(seven_days_ago_date, datetime.min.time())
    recent_notes_count = (
        db.query(func.count(models.Note.id))
        .filter(models.Note.created_at >= seven_days_ago)
        .scalar()
        or 0
    )

    return {
        "status_counts": status_counts,
        "type_counts": type_counts,
        "recent_notes_count": recent_notes_count,
    }


@app.get("/api/reminders", response_model=list[schemas.ReminderOut])
def list_reminders(db: Session = Depends(get_db)):
    """获取全部提醒。"""
    return (
        db.query(models.Reminder)
        .order_by(models.Reminder.completed.asc(), models.Reminder.reminder_date.asc())
        .all()
    )


@app.post("/api/reminders", response_model=schemas.ReminderOut, status_code=201)
def create_reminder(payload: schemas.ReminderCreate, db: Session = Depends(get_db)):
    """创建提醒。"""
    batch = db.query(models.Batch).filter(models.Batch.id == payload.batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")

    reminder = models.Reminder(**payload.model_dump())
    db.add(reminder)
    db.commit()
    db.refresh(reminder)
    return reminder


@app.get("/api/reminders/{reminder_id}", response_model=schemas.ReminderOut)
def get_reminder(reminder_id: int, db: Session = Depends(get_db)):
    """获取提醒详情。"""
    reminder = db.query(models.Reminder).filter(models.Reminder.id == reminder_id).first()
    if not reminder:
        raise HTTPException(status_code=404, detail="提醒不存在")
    return reminder


@app.put("/api/reminders/{reminder_id}", response_model=schemas.ReminderOut)
def update_reminder(
    reminder_id: int,
    payload: schemas.ReminderUpdate,
    db: Session = Depends(get_db),
):
    """更新提醒。"""
    reminder = db.query(models.Reminder).filter(models.Reminder.id == reminder_id).first()
    if not reminder:
        raise HTTPException(status_code=404, detail="提醒不存在")

    update_data = payload.model_dump(exclude_unset=True)
    if "batch_id" in update_data:
        batch = db.query(models.Batch).filter(models.Batch.id == update_data["batch_id"]).first()
        if not batch:
            raise HTTPException(status_code=404, detail="批次不存在")

    for key, value in update_data.items():
        setattr(reminder, key, value)

    db.commit()
    db.refresh(reminder)
    return reminder


@app.patch("/api/reminders/{reminder_id}/toggle", response_model=schemas.ReminderOut)
def toggle_reminder_completed(reminder_id: int, db: Session = Depends(get_db)):
    """切换提醒完成状态。"""
    reminder = db.query(models.Reminder).filter(models.Reminder.id == reminder_id).first()
    if not reminder:
        raise HTTPException(status_code=404, detail="提醒不存在")

    reminder.completed = not reminder.completed
    db.commit()
    db.refresh(reminder)
    return reminder


@app.delete("/api/reminders/{reminder_id}", status_code=204)
def delete_reminder(reminder_id: int, db: Session = Depends(get_db)):
    """删除提醒。"""
    reminder = db.query(models.Reminder).filter(models.Reminder.id == reminder_id).first()
    if not reminder:
        raise HTTPException(status_code=404, detail="提醒不存在")
    db.delete(reminder)
    db.commit()
